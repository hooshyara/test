from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from django.shortcuts import render
from .models import *
from .serializers import *
from permissoin.models import Permission
from user.helper import get_user
from django.db.models import Q
from django.db import transaction
import json
import weasyprint
from django.template.loader import get_template
from company.serializers import CompanySeriallizer
import jdatetime
from rest_framework.pagination import PageNumberPagination
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
class PaginationCustom(PageNumberPagination):
    page_size = 2  
    page_size_query_param = 'page_size'  
    max_page_size = 100 
    
# class ExportUsersToExcel(APIView):
#     def get(self, request):
#         # دریافت داده‌های کاربران از دیتابیس
#         users = User.objects.all()
        
#         # ایجاد یک workbook جدید
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Users Data"
        
#         # ایجاد هدرهای جدول
#         headers = [
#             "ID", "Username", "Full Name", 
#             "Phone", "Personal Code", "Company"
#         ]
        
#         for col_num, header in enumerate(headers, 1):
#             col_letter = get_column_letter(col_num)
#             ws[f"{col_letter}1"] = header
#             ws.column_dimensions[col_letter].width = 20
        
#         # پر کردن داده‌ها
#         for row_num, user in enumerate(users, 2):
#             ws[f"A{row_num}"] = user.id
#             ws[f"B{row_num}"] = user.username
#             ws[f"C{row_num}"] = user.fullName
#             ws[f"D{row_num}"] = user.phone
#             ws[f"E{row_num}"] = user.personal_code
#             ws[f"F{row_num}"] = str(user.company) if user.company else ""
        
#         # ایجاد پاسخ HTTP با فایل اکسل
#         response = HttpResponse(
#             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#         response["Content-Disposition"] = "attachment; filename=users_data.xlsx"
        
#         # ذخیره workbook در response
#         wb.save(response)
        
#         return response
class SailingView(APIView):
    def get(self, request):
        try:
            sailing = Sailing.objects.all()
            serializer = SailingSerializer(sailing, many=True)
            return Response({"data":
                            serializer.data}, status=status.HTTP_200_OK)
        except Sailing.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                sailing = Sailing.objects.create(
                    title=request.data.get('title'),
                    name=request.data.get('name'),
  
                )
                sailing.save()
                return Response({"Message": "OK"}, status=status.HTTP_201_CREATED)
            except Exception as e:
                print(e)
                return Response({"Message": "Error"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"Message": "You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
            
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                sailing = Sailing.objects.get(id=id)
                serializer = SailingSerializer(sailing, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            except :
                return Response({'message': 'your Target doesn\'t exist'}, 
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_list or permission.delete_item:
            try:
                sailing = Sailing.objects.get(id=id)
                sailing.delete()
                return Response({"message": "OK"}, status=status.HTTP_200_OK)
            
            except Exception as e:
                print(e)
                return Response({'message': 'your Target doesn\'t exist'},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
class OwnersOfGoodsView(APIView): # صاحبان کالا

    def get(self, request):
        try:
            owners_of_goods = OwnersOfGoods.objects.all()
            serializers = OwnersOfGoodsSeriallizer(owners_of_goods, many=True)
            return Response({"Message":serializers.data},
                             status=status.HTTP_200_OK)
        
        except:
            return Response({"Message":"'your Target doesn\'t exist'"},
                             status=status.HTTP_400_BAD_REQUEST)


    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_owner_of_goods and permission.add_list_item:
            owners_of_goods = OwnersOfGoods.objects.create(
                name=request.data.get('name'),
                tel=request.data.get('tel'),
                fax=request.data.get('fax'),
                mobile=request.data.get('mobile'),
                address=request.data.get('address'),
                commission=request.data.get('commission'),
            )
            owners_of_goods.save()
            return Response({"message": 
                                "OK"}, status=status.HTTP_200_OK)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_owner_of_goods :
            try:
                token = request.data.get("token")
                user = get_user(token)
                permission = Permission.objects.get(user=user)
                
                owners_of_goods = OwnersOfGoods.objects.get(id=id)
                serializers = OwnersOfGoodsSeriallizer(owners_of_goods,
                                            data=request.data, partial=True)
                if serializers.is_valid():
                    serializers.save()
                    return Response({"message": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"ERROR": serializers.errors},
                                    status=status.HTTP_400_BAD_REQUEST)
            except:
                return Response({'message': 'your Target doesn\'t exist'},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)


    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_owner_of_goods :
            try:
                owners_of_goods = OwnersOfGoods.objects.get(id=id)
                owners_of_goods.delete()
                return Response({"message": "OK"}, status=status.HTTP_200_OK)
            except:
                return Response({'message': 'your Target doesn\'t exist'},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
class BillDescribtionView(APIView): # مشخصات بارنامه
    def get(self, request, id=None, name=None):
        if id:
            bill = BillDescribtion.objects.get(id=id)

            continer = Continer.objects.filter(bill=bill)
            continer_id = continer.values_list("file_number", flat=True)
            print(continer_id)
            cont_ser = ContinerSerializers(continer, many=True)

            guides = Guide.objects.filter(Q(continer__file_number__in=continer_id), Q(is_valid=True))
            guide_ser = GuideSerializers(guides, many=True)
            guide_id = guides.values_list("id", flat=True)
            booking = Booking.objects.filter(guide__id__in=guide_id)
            booking_ser = BookingSerializers(booking, many=True)
            
            return Response({"continer":cont_ser.data, 
                            "booking":booking_ser.data, "guide":guide_ser.data}, 
                            status=status.HTTP_200_OK)
        elif name:
            paginator = PageNumberPagination()
            paginator.page_size = request.GET.get('page_size', 2)
            
            company = MainCompany.objects.get(main_company__icontains=name)
            bill = BillDescribtion.objects.filter(company=company)
            bill_page = paginator.paginate_queryset(bill, request) 
            bill_ser = BillDescribtionSeriallizer(bill_page, many=True)
            return paginator.get_paginated_response({"data":bill_ser.data})
        else:
            try:
                continer = Continer.objects.all()
                continer_count = len(continer)
                demurrage_type = DemurrageType.objects.all()
                demurrage_type_serializers = DemurrageTypeSerializers(demurrage_type, many=True)
                get_documents_type = GetDocumentType.objects.all()
                get_documents_type_serializers = GetDocumentTypeSerializers(get_documents_type, many=True)
                bill_describtions = BillDescribtion.objects.all()
                serializers = BillDescribtionSeriallizer(bill_describtions, many=True)
                sailing = Sailing.objects.all()
                sailing_serializers = SailingSerializer(sailing, many=True)
                owners_of_goods = OwnersOfGoods.objects.all()
                owners_of_goods_serializers = OwnersOfGoodsSeriallizer(owners_of_goods,
                                                                        many=True)
                return Response({"Message":{"bill":serializers.data,
                "continer_count":continer_count,
                "ship":sailing_serializers.data,
                "demurrage_type":demurrage_type_serializers.data,
                "get_documents_type":get_documents_type_serializers.data,
                "owners_of_goods":owners_of_goods_serializers.data}},
                    status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"'your Target doesn\'t exist'"},
                                status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        token = request.data.get('token')
        user = get_user(token)
        company= MainCompany.objects.get(id=request.data.get('company'))
        permission = Permission.objects.get(user=user)
        if permission.add_bill and permission.add_list_item:
            try:
                get_documents_type = GetDocumentType.objects.get(id=request.data.get('get_documents_type'))
                demurrage_type = DemurrageType.objects.get(id=request.data.get('demurrage_type'))
                sailings = Sailing.objects.get(id=request.data.get('ship'))
                owners_of_goods = OwnersOfGoods.objects.get(id=request.data.get('Owners_of_goods'))
                bill_describtions = BillDescribtion.objects.create(
                    get_documents_type = get_documents_type,
                    bill_lading_number = request.data.get('bill_lading_number'),
                    ship = sailings,
                    Owners_of_goods = owners_of_goods,
                    Free_Time = request.data.get('Free_Time'),
                    demurrage_type = demurrage_type,
                    containers_count = request.data.get('containers_count'),
                    total_weight_load = request.data.get('total_weight_load'),
                    perweight = request.data.get('perweight'),
                    total_product_count = request.data.get('total_product_count'),
                    per_product_count = request.data.get('per_product_count'),
                    Reference_code = request.data.get('Reference_code'),
                    ship_name = request.data.get('ship_name'),
                    company = company
                )
                bill_describtions.save()
                return Response({"message": "OK"}, status=status.HTTP_200_OK)
            
            except Exception as e:
                print(e)
                return Response({"message": "Error"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get('token')
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_bill and permission.update_item:
            try:
                bill_describtions = BillDescribtion.objects.get(id=id)
                bill_describtions = BillDescribtionSeriallizer(bill_describtions, data=request.data,
                                                            partial=True)
                if bill_describtions.is_valid():
                    bill_describtions.save()
                    return Response({"message": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"message": "Error"}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"message": "Error"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

    def delete(self, request, id):
        try:
            bill_describtions = BillDescribtion.objects.get(id=id)
            bill_describtions.delete()
            return Response({"message": "OK"}, status=status.HTTP_200_OK)
        except:
            return Response({"message": "your Target doesn't exist"},
                             status=status.HTTP_404_NOT_FOUND)
        
class RouteView(APIView):
    def get(self, request, id=None):
        try:
            city = City.objects.all()
            city_serializers = Cityserializers(city, many=True,
                                            context={'request': request})
            route = Route.objects.all()
            serializers = RouteSerializers(route, many=True)
            country = Country.objects.all()
            country_serializers = Countryserializers(country, many=True,
                                                    context={'request': request})
            return Response({"route": serializers.data, "city":city_serializers.data,
                            "country":country_serializers.data},
                            status=status.HTTP_200_OK)
        except:
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_400_BAD_REQUEST),

    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        
        permission = Permission.objects.get(user=user)
        if permission.add_rout and permission.add_list_item:
            try:
                route = Route.objects.create(
                    source_country=request.data.get('source_country'),
                    source_city=request.data.get('source_city'),
                    distinction_country=request.data.get('distinction_country'),
                    distinction_city=request.data.get('distinction_city'),
                    Insuranceـpremium=request.data.get('Insuranceـpremium'),
                    Driver_exemption=request.data.get('Driver_exemption'),
                    Price_per_ton_kilometer=request.data.get('Price_per_ton_kilometer'),
                    Container_departure_price=request.data.get('Container_departure_price'),
                )
                route.save()
                route.title = route.source_city + " - "+route.distinction_city
                route.save()
                return Response({"MESSAGE": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_400_BAD_REQUEST),

        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_rout and permission.update_item:
            try:
                route = Route.objects.get(id=id)
                serializers = RouteSerializers(route, data=request.data, partial=True)
                if serializers.is_valid():
                    serializers.save()
                    return Response({"data": serializers.data}, 
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"message": serializers.errors},
                                    status=status.HTTP_400_BAD_REQUEST)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_rout and permission.delete_item:
            try:
                route = Route.objects.get(id=id)
                route.delete()
                return Response({"message": "OK"}, status=status.HTTP_200_OK)
            except:
                return Response({"message": "your Target doesn't exist"}),
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class StationView(APIView):
    def get(self, request, id=None):
        try:
            route = Route.objects.all()
            station = Station.objects.all()
            route_serializers = RouteSerializers(route, many=True)
            serializers = StationSerializers(station, many=True)
            return Response({"station": serializers.data, "route":route_serializers.data}
                            , status=status.HTTP_200_OK)
        except:
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_station and permission.add_list_item:
            try:
                route = Route.objects.get(id=request.data.
                                                    get('route'))
                station = Station.objects.create(
                    name=request.data.get('name'),
                    route=route,
                )
                station.save()
                print(route)
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_station and permission.update_item:
            try:
                station = Station.objects.get(id=id)
                station_serializers = StationSerializers(station, data=request.data,
                                                        partial=True)
                if station_serializers.is_valid():
                    station_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"message": "your Target doesn't exist"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_station and permission.delete_item:
            try:
                station = Station.objects.get(id=id)
                station.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class DriverView(APIView):
    def get(self, request, id=None):
        try:
            driver = Driver.objects.all()
            country = Country.objects.all()
            country_ser = Countryserializers(country, many=True)
            driver_serializers = DriverSerializers(driver, many=True)
            return Response({"driver": driver_serializers.data,
                            "country":country_ser.data},status=status.HTTP_200_OK)
        except:
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
    
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_driver and permission.add_list_item:
            try:
                driver = Driver.objects.create(
                    name=request.data.get('name'),
                    country=request.data.get('country'),
                    national_id=request.data.get('national_id'),
                    tel=request.data.get('tel'),
                    mobile=request.data.get('mobile'),
                    smart_cart_number=request.data.get('smart_cart_number'),
                    cart_number=request.data.get('cart_number'),
                    cart_validate=request.data.get('cart_validate'),
                    Booklet_number=request.data.get('Booklet_number'),
                    Booklet_validate=request.data.get('Booklet_validate'),
                    Certificate_number=request.data.get('Certificate_number'),
                    Certificate_validate=request.data.get('Certificate_validate'),
                    address=request.data.get('address'),
                )
                driver.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_driver and permission.update_item:
            try:
                driver = Driver.objects.get(id=id)
                driver_serializers = DriverSerializers(driver, data=request.data,
                                                        partial=True)
                if driver_serializers.is_valid():
                    driver_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"message": "your Target doesn't exist"},
                                    status=status.HTTP_404_NOT_FOUND)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_driver and permission.delete_item:
            try:
                driver = Driver.objects.get(id=id)
                driver.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class InsuranceView(APIView): #بیمه
    def get(self, request):
        try:
            insurance = Insurance.objects.all()
            insurance_serializers = InsuranceSerializers(insurance, many=True)
            return Response({"data": insurance_serializers.data}, status=status.HTTP_200_OK)
        except:
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
    
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                insurance = Insurance.objects.create(
                    name=request.data.get('name'),
                    title=request.data.get('title'),
                )
                insurance.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
            
    def put(self ,request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                insurance = Insurance.objects.get(id=id)
                insurance_serializers = InsuranceSerializers(insurance, data=request.data,
                                                            partial=True)
                if insurance_serializers.is_valid():
                    insurance_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"message": "your Target doesn't exist"},
                                    status=status.HTTP_404_NOT_FOUND)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_list or permission.delete_item:
            try:
                insurance = Insurance.objects.get(id=id)
                insurance.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

class TruckTypeView(APIView):
    def get(self, request):
        try:
            truck_type = TruckType.objects.all()
            truck_type_serializers = TruckTypeSerializers(truck_type, many=True)
            return Response({"data": truck_type_serializers.data},
                             status=status.HTTP_200_OK)
        except:
            return Response({"Message":"ERROR"},
                             status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                truck_type = TruckType.objects.create(
                    name=request.data.get('name'),
                    title=request.data.get('title'),
                    )
                truck_type.save()
                return Response({"Message": "OK"},
                                status=status.HTTP_200_OK)
            except:
                return Response({"Message": "ERROR"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"Message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                truck_type = TruckType.objects.get(id=id)
                truck_type_serializers = TruckTypeSerializers(truck_type, data=request.data,
                                                            partial=True)
                if truck_type_serializers.is_valid():
                    truck_type_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"message": truck_type_serializers.errors},
                                    status=status.HTTP_404_NOT_FOUND)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                truck_type = TruckType.objects.get(id=id)
                truck_type.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
class TruckView(APIView):
    def get(self, request):
        try:
            trailer = Trailer.objects.all()
            trailer_serializers = TrailerSerializers(trailer, many=True,
                                                      context={'request': request})
            truck_type = TruckType.objects.all()
            truck_type_serializers = TruckTypeSerializers(truck_type,
                                    many=True, context={'request': request})
            country = Country.objects.all()
            country_serializers = Countryserializers(country, many=True)
            truck = Truck.objects.all()
            truck_serializers = TruckSerializers(truck, many=True)
            return Response({"truck": truck_serializers.data,
                            "truck_type":truck_type_serializers.data,
                            "country":country_serializers.data,
                            "trailer":trailer_serializers.data}, 
                            status=status.HTTP_200_OK)
        except:
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get('token')
        user = get_user(token)
        print(user)
        permission = Permission.objects.get(user=user)
        if permission.add_truck and permission.add_list_item:
            try:
                truck_type = TruckType.objects.get(id=request.data.get('truck_type'))
                trailer= Trailer.objects.get(id=request.data.get('trailer'))
                country = Country.objects.get(id=request.data.get('country'))
                truck = Truck.objects.create(
                    truck_type=truck_type,
                    trailer= trailer,
                    country=country,
                    License_plate_number=request.data.get('License_plate_number'),
                    Smart_cart_number=request.data.get('Smart_cart_number'),
                    Transit_number=request.data.get('Transit_number'),
                    address=request.data.get('address'),
                    insurance_number=request.data.get('insurance_number'),
                    license_number=request.data.get('license_number'),
                    Hood_number=request.data.get('Hood_number'),
                    date_issuance_hood=request.data.get('date_issuance_hood'),
                    date_validity_hood=request.data.get('date_validity_hood'),
                    owner_name=request.data.get('owner_name'),
                    mobile=request.data.get('mobile'),
                    Annual_contract_number=request.data.get('Annual_contract_number'),
                )
                truck.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"MESSAGE": "ERROR"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"MESSAGE": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get('token')
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_truck and permission.update_item:
            try:
                truck = Truck.objects.get(id=id)
                truck_serializers = TruckSerializers(truck, data=request.data,
                                                            partial=True)
                if truck_serializers.is_valid():
                    truck_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"MESSAGE": truck_serializers.errors},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"MESSAGE": "ERROR"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"MESSAGE": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get('token')
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_truck and permission.delete_item:
            try:
                truck = Truck.objects.get(id=id)
                truck.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"MESSAGE": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class TrailerView(APIView):
    def get(self, request):
        try:
            
            trailer = Trailer.objects.all()
            trailer_serializers = TrailerSerializers(trailer, many=True)
            return Response({"data": trailer_serializers.data}, 
                            status=status.HTTP_200_OK)
        except:
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
    
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                trailer = Trailer.objects.create(
                    name = request.data.get('name'),
                    title = request.data.get('title'),
                )
                trailer.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"MESSAGE": "ERROR"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)   
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                trailer = Trailer.objects.get(id=id)
                trailer_serializers = TrailerSerializers(trailer, data=request.data,
                                                        partial=True)
                if trailer_serializers.is_valid():
                    trailer_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"MESSAGE": trailer_serializers.errors},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                trailer = Trailer.objects.get(id=id)
                trailer.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class CarrierView(APIView):
    def get(self, request):
        try:
            carrier = Carrier.objects.all()
            carrier_serializers = CarrierSerializers(carrier, many=True)
            return Response({"data": carrier_serializers.data}, status=status.HTTP_200_OK)
        except:
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        print(user)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                carrier = Carrier.objects.create(
                    name=request.data.get('name'),
                    title=request.data.get('title'), 
                    
                )
                carrier.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_list or permission.update_item:
            try:
                carrier = Carrier.objects.get(id=id)
                carrier_serializers = CarrierSerializers(carrier, data=request.data,
                                                        partial=True)
                if carrier_serializers.is_valid():
                    carrier_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"message": "your Target doesn't exist"},
                                    status=status.HTTP_404_NOT_FOUND)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                carrier = Carrier.objects.get(id=id)
                carrier.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except:
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class PackagingTypeView(APIView):
    def get(self, request):
        try:
            packaging_type = PackagingType.objects.all()
            packaging_type_serializers = PackagingTypeSerializers(packaging_type,
                                                                many=True)
            return Response({"data": packaging_type_serializers.data},
                             status=status.HTTP_200_OK)
        except:
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                packaging_type = PackagingType.objects.create(
                    name=request.data.get('name'),
                    title=request.data.get('title'),
                    )
                packaging_type.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
    
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                packaging_type = PackagingType.objects.get(id=id)
                packaging_type_serializers = PackagingTypeSerializers(
                    packaging_type, data=request.data, partial=True)
                if packaging_type_serializers.is_valid():
                    packaging_type_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"message": "your Target doesn't exist"},
                                    status=status.HTTP_404_NOT_FOUND)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
    
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                packaging_type = PackagingType.objects.get(id=id)
                packaging_type.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
    
class DeliveryRecipientCMRView(APIView):
    def get(self, request):
        try:
            delivery_recipient = DeliveryRecipientCMR.objects.all()
            delivery_recipient_serializers = DeliveryRecipientCMRSerializers(delivery_recipient, 
                                        many=True, context={'request': request})
            return Response({"data":delivery_recipient_serializers.data},
                             status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)

    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                delivery_recipient = DeliveryRecipientCMR.objects.create(
                    name=request.data.get('name'),
                    title=request.data.get('title'),
                )
                delivery_recipient.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        

    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                delivery_recipient = DeliveryRecipientCMR.objects.get(id=id)
                delivery_recipient_serializers = DeliveryRecipientCMRSerializers(delivery_recipient,
                    data=request.data, context={'request': request})
                if delivery_recipient_serializers.is_valid():
                    delivery_recipient_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
            
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                delivery_recipient = DeliveryRecipientCMR.objects.get(id=id)
                delivery_recipient.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

class DeliveryRecipientProductView(APIView):
    def get(self, request):
        try:
            delivery_recipient_product = DeliveryRecipientProduct.objects.all()
            delivery_recipient_product_serializers = DeliveryRecipientProductSerializers(delivery_recipient_product, many
                                                                                         =True, context={'request': request})
            return Response({"data":delivery_recipient_product_serializers.data},
                             status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                delivery_recipient_product = DeliveryRecipientProduct.objects.create(
                    name=request.data.get('name'),
                    title=request.data.get('title'), 
                )
                delivery_recipient_product.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
    
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                delivery_recipient_product = DeliveryRecipientProduct.objects.get(id=id)
                delivery_recipient_product_serializers = DeliveryRecipientProductSerializers(delivery_recipient_product,
                                    data=request.data,context={'request': request})
                if delivery_recipient_product_serializers.is_valid():
                    delivery_recipient_product_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response(delivery_recipient_product_serializers.errors,
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                delivery_recipient_product = DeliveryRecipientProduct.objects.get(id=id)
                delivery_recipient_product.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

class DeliveryRecipientBillView(APIView):
    def get(self, request):
        try:
            delivery_recipient_bill = DeliveryRecipientBill.objects.all()
            delivery_recipient_bill_serializers = DeliveryRecipientBillSerializers(delivery_recipient_bill,
                            many= True, context={'request': request})
            return Response({"data":delivery_recipient_bill_serializers.data},
                             status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)

    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                delivery_recipient_bill = DeliveryRecipientBill.objects.create(
                    name = request.data.get('name'),
                    title = request.data.get('title'),
                )
                delivery_recipient_bill.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                delivery_recipient_bill = DeliveryRecipientBill.objects.get(id=id)
                delivery_recipient_bill_serializers = DeliveryRecipientBillSerializers(delivery_recipient_bill,
                                                                                    data=request.data,context={'request': request})
                if delivery_recipient_bill_serializers.is_valid():
                    delivery_recipient_bill_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response(delivery_recipient_bill_serializers.errors,
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                delivery_recipient_bill = DeliveryRecipientBill.objects.get(id=id)
                delivery_recipient_bill.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class CityView(APIView):
    def get(self, request):
        try:
            city = City.objects.all()
            city_serializers = Cityserializers(city, many=True)
            return Response({"data": city_serializers.data},
                             status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                city = City.objects.create(
                    name=request.data.get('name'),
                    title=request.data.get('title'),
                )
                city.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_list or permission.update_item:
            try:
                city = City.objects.get(id=id)
                city_serializers = Cityserializers(city, data=request.data, partial=True)
                if city_serializers.is_valid():
                    city_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"ERROR":city_serializers.errors},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                city = City.objects.get(id=id)
                city.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class CountryView(APIView):
    def get(self, request):
        try:
            country = Country.objects.all()
            country_serializers = Countryserializers(country, many=True)
            return Response({"data": country_serializers.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
    
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                country = Country.objects.create(
                    name=request.data.get('name'),
                    title=request.data.get('title'),
                )
                country.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                country = Country.objects.get(id=id)
                country_serializers = Countryserializers(country, data=request.data, partial=True)
                if country_serializers.is_valid():
                    country_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"message": "your Target doesn't exist"},
                                    status=status.HTTP_404_NOT_FOUND)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                country = Country.objects.get(id=id)
                country.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class ProductView(APIView):
    def get(self, request):
        try:
            product = Product.objects.all()
            product_serializers = Productserializers(product, many=True)
            return Response({"data": product_serializers.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
    
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                product = Product.objects.create(
                    name=request.data.get('name'),
                    title=request.data.get('title'),
                    )
                product.save()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message": "You Don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                product = Product.objects.get(id=id)
                product_serializers = Productserializers(product, data=request.data, partial=True)
                if product_serializers.is_valid():
                    product_serializers.save()
                    return Response({"data": "OK"}, status=status.HTTP_200_OK)
                else:
                    return Response({"message": "your Target doesn't exist"},
                                    status=status.HTTP_404_NOT_FOUND)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message" : "You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                product = Product.objects.get(id=id)
                product.delete()
                return Response({"data": "OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message" : "You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
class CreateBookingView(APIView): # get before put booking
    def get(self, request, id):
        try:
            carrier= Carrier.objects.all()
            carrier_serializer = CarrierSerializers(carrier, many=True)
            route =Route.objects.all()
            serializer_route = RouteSerializers(route, many=True)
            truck =Truck.objects.all()
            serializer_truck = TruckSerializers(truck, many=True)
            company_info =Company.objects.all()
            serializer_company = CompanySeriallizer(company_info, many=True)
            driver =Driver.objects.all()
            serializer_driver = DriverSerializers(driver, many=True)
            owners_of_goods =OwnersOfGoods.objects.all()
            serializer_owners_of_goods = OwnersOfGoodsSeriallizer(owners_of_goods, many=True)
            continer = Continer.objects.get(file_number=id)
            guide =Guide.objects.get(continer=continer.file_number)
            serializer_guide = GuideSerializers(guide, many=False)
            serializer_continer = ContinerSerializers(continer, many=False)
            booking = Booking.objects.filter(continer_number=continer)
            booking_serializers = BookingSerializers(booking, many=True)
            return Response({"booking": booking_serializers.data,
                            "carrier":carrier_serializer.data,
                            "route":serializer_route.data, 
                            "truck":serializer_truck.data, 
                            "company":serializer_company.data, 
                            "driver":serializer_driver.data, 
                            "owner_of_good":serializer_owners_of_goods.data, 
                            "guide":serializer_guide.data, 
                            "continer":serializer_continer.data}, 
                            status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
class BookingView(APIView):
    def get(self, request, id):
        try:
            carrier= Carrier.objects.all()
            carrier_serializer = CarrierSerializers(carrier, many=True)
            route =Route.objects.all()
            serializer_route = RouteSerializers(route, many=True)
            truck =Truck.objects.all()
            serializer_truck = TruckSerializers(truck, many=True)
            company_info =Company.objects.all()
            serializer_company = CompanySeriallizer(company_info, many=True)
            driver =Driver.objects.all()
            serializer_driver = DriverSerializers(driver, many=True)
            owners_of_goods =OwnersOfGoods.objects.all()
            serializer_owners_of_goods = OwnersOfGoodsSeriallizer(owners_of_goods, many=True)
            guide =Guide.objects.filter(company=id)
            
            serializer_guide = GuideSerializers(guide, many=True)
            continer = Continer.objects.all()
            serializer_continer = ContinerSerializers(continer, many=True)
            booking = Booking.objects.filter(company=id)
            # cont = Continer.objects.get()
            booking_serializers = BookingSerializers(booking, many=True)
            return Response({"booking": booking_serializers.data,
                            "carrier":carrier_serializer.data,
                            "route":serializer_route.data, 
                            "truck":serializer_truck.data, 
                            "company":serializer_company.data, 
                            "driver":serializer_driver.data, 
                            "owner_of_good":serializer_owners_of_goods.data, 
                            "guide":serializer_guide.data, 
                            "continer":serializer_continer.data}, 
                            status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                print(f' req : {request.data}')
                carrier = Carrier.objects.get(id=request.data.get('carrier'))
                route = Route.objects.get(id=request.data.get('route'))
                truck = Truck.objects.get(id=request.data.get('truck'))
                company_info = Company.objects.get(id=request.data.get('company_info'))
                driver = Driver.objects.get(id=request.data.get('driver'))
                print(request.data.get('continer_number'))
                owners_of_goods = OwnersOfGoods.objects.get(id=request.data.get('owners_of_goods'))
                company= MainCompany.objects.get(id=request.data.get('company'))
                continer = Continer.objects.get(file_number=request.data.get('continer_number'))
                guide = Guide.objects.get(continer=continer.file_number)
                booking = Booking.objects.create(
                    carrier=carrier,
                    route=route,
                    truck=truck,
                    company_info=company_info,
                    driver=driver,
                    guide=guide,
                    owners_of_goods=owners_of_goods,
                    company=company,
                    booking_number = request.data.get('booking_number'),
                    continer_number = continer,
                    loading_date = request.data.get('loading_date'),
                    guide_num = request.data.get('guide_num'),
                    product_weight = request.data.get('product_weight'),
                    rent = request.data.get('rent'),
                    tonnageـweight = request.data.get('tonnageـweight'),
                    tonnage_count = request.data.get('tonnage_count'),
                    tonnageـamount = request.data.get('tonnageـamount'),
                    commission_type = request.data.get('commission_type'),
                    commission_percent = request.data.get('commission_percent'),
                    commission_amount = request.data.get('commission_amount'),
                    total_amount_owners_of_goods = request.data.get('total_amount_owners_of_goods'),
                    placedـnumber = request.data.get('placedـnumber'),
                    rent_driver = request.data.get('rent_driver'),
                    tonnage_count_driver = request.data.get('tonnage_count_driver'),
                    tonnageـamount_driver = request.data.get('tonnageـamount_driver'),
                    portـcommission = request.data.get('portـcommission'),
                    parking = request.data.get('parking'),
                    perrent_port = request.data.get('perrent_port'),
                    total_amount_driver = request.data.get('total_amount_driver'),
                    driver_accountـbalance = request.data.get('driver_accountـbalance'),
                    difference = request.data.get('difference'),
                    discribtion = request.data.get('discribtion'),
                )
                booking.save()
                log = Log.objects.create(
                    booking=booking,
                    company = company,
                    carrier=booking.carrier,
                    route=booking.route,
                    truck=booking.truck,
                    company_info=booking.company_info,
                    driver=booking.driver,
                    guide=booking.guide,
                    owners_of_goods=booking.owners_of_goods,
                    # booking_number = request.data.get('booking_number'),
                    continer_number = booking.continer_number,
                    loading_date = booking.loading_date,
                    guide_num = booking.guide_num,
                    product_weight = booking.product_weight,
                    rent = booking.rent,
                    tonnageـweight = booking.tonnageـweight,
                    tonnage_count = booking.tonnage_count,
                    tonnageـamount = booking.tonnageـamount,
                    commission_type = booking.commission_type,
                    commission_percent = booking.commission_percent,
                    commission_amount = booking.commission_amount,
                    total_amount_owners_of_goods = booking.total_amount_owners_of_goods,
                    placedـnumber = booking.placedـnumber,
                    rent_driver = booking.rent_driver,
                    tonnage_count_driver = booking.tonnage_count_driver,
                    tonnageـamount_driver = booking.tonnageـamount_driver,
                    portـcommission = booking.portـcommission,
                    parking = booking.parking,
                    perrent_port = booking.perrent_port,
                    total_amount_driver = booking.total_amount_driver,
                    driver_accountـbalance = booking.driver_accountـbalance,
                    difference = booking.difference,
                    discribtion = booking.discribtion,
                )
                log.save()
                return Response({"message": "Booking created successfully"},
                                status=status.HTTP_201_CREATED)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "you don;t have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                booking = Booking.objects.get(id=id)
                booking_serializers = BookingSerializers(booking,
                                                data=request.data, partial=True)
                if booking_serializers.is_valid():
                    booking_serializers.save()
                    log = Log.objects.get(booking=booking)
                    log.update_status = True
                    log.save()
                    return Response({"message": "Booking updated successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    print(f'error : {booking_serializers.errors}')
                    return Response({"message": "Invalid data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                booking = Booking.objects.get(id=id)
                booking.delete()
                return Response({"message": "Booking deleted successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

class BookingCreateView(APIView):
    def get(self, request, id):
        try:
            carrier= Carrier.objects.all()
            carrier_serializer = CarrierSerializers(carrier, many=True)
            route =Route.objects.all()
            serializer_route = RouteSerializers(route, many=True)
            truck =Truck.objects.all()
            serializer_truck = TruckSerializers(truck, many=True)
            company_info =Company.objects.all()
            serializer_company = CompanySeriallizer(company_info, many=True)
            driver =Driver.objects.all()
            serializer_driver = DriverSerializers(driver, many=True)
            owners_of_goods =OwnersOfGoods.objects.all()
            serializer_owners_of_goods = OwnersOfGoodsSeriallizer(owners_of_goods, many=True)
            
            return Response({
                            "carrier":carrier_serializer.data,
                            "route":serializer_route.data, 
                            "truck":serializer_truck.data, 
                            "company":serializer_company.data, 
                            "driver":serializer_driver.data, 
                            "owner_of_good":serializer_owners_of_goods.data, 
                            }, 
                            status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)


class GetDocumentTypeView(APIView):
    def get(self, request):
        try:
            document_type = GetDocumentType.objects.all()
            serializers = GetDocumentTypeSerializers(document_type,
                                many=True, context={'request': request})
            return Response({"data": serializers.data},
                            status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                document_type = GetDocumentType.objects.create(
                    name = request.data.get('name'),
                    title = request.data.get('title'), 
                )
                document_type.save()
                return Response({"Message":"OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
         
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                document_type = GetDocumentType.objects.get(id=id)
                document_type_serializers = GetDocumentTypeSerializers(document_type,
                                    data=request.data, context={'request': request}, partial=True)
                if document_type_serializers.is_valid():
                    document_type_serializers.save()
                    return Response({"message": "Document Type updated successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    print(document_type_serializers.errors)
                    return Response({"message": "Invalid data"},
                                    status=status.HTTP_400_BAD_REQUEST)
                
            except Exception as e:
                print(e)
                return Response({"message": "Invalid data"},
                                    status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
                
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                document_type = GetDocumentType.objects.get(id=id)
                document_type.delete()
                return Response({"message": "Document Type deleted successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "Invalid data"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class EnterTypeView(APIView):
    def get(self, request):
        try:
            enter_type = EnterType.objects.all()
            enter_type_serializers = EnterTypeSerializers(enter_type, many=True)
            return Response({"data": enter_type_serializers.data},
                            status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                enter_type = EnterType.objects.create(
                    name = request.data.get('name'),
                    title = request.data.get('title'), 
                )
                enter_type.save()
                return Response({"Message":"OK"}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                enter_type = EnterType.objects.get(id=id)
                enter_type_serializers = EnterTypeSerializers(enter_type, 
                                                            data=request.data, partial=True)
                if enter_type_serializers.is_valid():
                    enter_type_serializers.save()
                    return Response({"message": "Enter Type updated successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"message": "Invalid data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"message": "Invalid data"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
 
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                enter_type = EnterType.objects.get(id=id)
                enter_type.delete()
                return Response({"message": "Enter Type deleted successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": "Invalid data"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class PostalCompanyView(APIView):
    def get(self, request):
        try:
            postal_company = PostalCompany.objects.all()
            postal_company_serializers = PostalCompanySerializers(postal_company,
                                        many=True)
            return Response({"data":postal_company_serializers.data},
                            status=status.HTTP_200_OK)
        
        except Exception as e:
            print(e)
            return Response({"message": "your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                postal_company = PostalCompany.objects.create(
                    title = request.data.get('title'),
                    name = request.data.get('name'),
                )
                postal_company.save()
                return Response({"message": "Postal Company created successfully"},
                                status=status.HTTP_200_OK)
            
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
            
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                postal_company = PostalCompany.objects.get(id=id)
                postal_company_serializers = PostalCompanySerializers(postal_company,
                                                                    data=request.data, partial=True)
                if postal_company_serializers.is_valid():
                    postal_company_serializers.save()
                    return Response({"message": "Postal Company updated successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    print(postal_company_serializers.errors)
                    return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
            except Exception as e :
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                postal_company = PostalCompany.objects.get(id=id)
                postal_company.delete()
                return Response({"message": "Enter Type deleted successfully"},
                                status=status.HTTP_200_OK)
            
            except Exception as e:
                print(e)
                return Response({"message": "Invalid data"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class CostListView(APIView):
    def get(self, request):
        try:
            cost = CostList.objects.all()
            cost_serializers = CostListSerializers(cost, many=True)
            return Response({"data":cost_serializers.data},
                            status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                cost = CostList.objects.create(
                    title = request.data.get('title'),
                    name = request.data.get('name'),
                    company =None
                )
                cost.save()
                return Response({"message": "Cost List created successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                cost = CostList.objects.get(id=id)
                cost_serializers = CostListSerializers(cost, data=request.data, partial=True)
                if cost_serializers.is_valid():
                    cost_serializers.save()
                    return Response({"Message":"Update successfully"},
                                status=status.HTTP_200_OK)
            
            except Exception as e :
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
            
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                cost = CostList.objects.get(id=id)
                cost.delete()
                return Response({"message": "Enter Type deleted successfully"},
                                status=status.HTTP_200_OK)
            
            except Exception as e:
                print(e)
                return Response({"message": "Invalid data"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
            
class DemurrageTypeView(APIView):
    def get(self, request):
        try:
            demurrage = DemurrageType.objects.all()
            demurrage_serializers = DemurrageTypeSerializers(demurrage, many=True
                                                             )
            return Response({"data":demurrage_serializers.data},
                             status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                demurrage = DemurrageType.objects.create(
                    title = request.data.get('title'),
                    name = request.data.get('name'), 
                )
                demurrage.save()
                return Response({"Message":"Create successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                demurrage = DemurrageType.objects.get(id=id)
                demurrage_ser = DemurrageTypeSerializers(demurrage, data=request.data, partial=True)
                if demurrage_ser.is_valid():
                    demurrage_ser.save()
                    return Response({"Message":"Update successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"Message": "Invalid data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_list or permission.delete_item:
            try:
                demurrage = DemurrageType.objects.get(id=id)
                demurrage.delete()
                return Response({"Message":"Delete successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class CmrSalamnView(APIView):
    def get(self, request):
        try:
            cmr = CmrSalamn.objects.all()
            cmr_ser = CmrSalamnSerializers(cmr, many=True)
            return Response({"data":cmr_ser.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
            
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                cmr = CmrSalamn.objects.create(
                    title = request.data.get('title'),
                    name = request.data.get('name'),
                )
                cmr.save()
                return Response({"Message":"Create successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                cmr = CmrSalamn.objects.get(id=id)
                cmr_ser = CmrSalamnSerializers(cmr, data=request.data, partial=True)
                if cmr_ser.is_valid():
                    cmr_ser.save()
                    return Response({"Message":"Update successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"Message":"Invalid Data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                cmr = CmrSalamn.objects.get(id=id)
                cmr.delete()
                return Response({"Message":"Delete successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class CmrBarjameView(APIView):
    def get(self, request):
        try:
            cmr = CmrBarjame.objects.all()
            cmr_ser = CmrBarjameSerializers(cmr, many=True)
            return Response({"data":cmr_ser.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
            
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                cmr = CmrBarjame.objects.create(
                    title = request.data.get('title'),
                    name = request.data.get('name'),
                )
                cmr.save()
                return Response({"Message":"Create successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                cmr = CmrBarjame.objects.get(id=id)
                cmr_ser = CmrBarjameSerializers(cmr, data=request.data, partial=True)
                if cmr_ser.is_valid():
                    cmr_ser.save()
                    return Response({"Message":"Update successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"Message":"Invalid Data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                cmr = CmrBarjame.objects.get(id=id)
                cmr.delete()
                return Response({"Message":"Delete successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class CmrFerdowsView(APIView):
    def get(self, request):
        try:
            cmr = CmrFerdows.objects.all()
            cmr_ser = CmrFerdowsserializers(cmr, many=True)
            return Response({"data":cmr_ser.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
            
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                cmr = CmrFerdows.objects.create(
                    title = request.data.get('title'),
                    name = request.data.get('name'),
                )
                cmr.save()
                return Response({"Message":"Create successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                cmr = CmrFerdows.objects.get(id=id)
                cmr_ser = CmrFerdowsserializers(cmr, data=request.data, partial=True)
                if cmr_ser.is_valid():
                    cmr_ser.save()
                    return Response({"Message":"Update successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"Message":"Invalid Data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_list or permission.delete_item:
            try:
                cmr = CmrFerdows.objects.get(id=id)
                cmr.delete()
                return Response({"Message":"Delete successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class CmrMashahirView(APIView):
    def get(self, request):
        try:
            cmr = CmrMashahir.objects.all()
            cmr_ser = CmrMashahirSerializers(cmr, many=True)
            return Response({"data":cmr_ser.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
            
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                cmr = CmrMashahir.objects.create(
                    title = request.data.get('title'),
                    name = request.data.get('name'),
                )
                cmr.save()
                return Response({"Message":"Create successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_item or permission.update_list:
            try:
                cmr = CmrMashahir.objects.get(id=id)
                cmr_ser = CmrMashahirSerializers(cmr, data=request.data, partial=True)
                if cmr_ser.is_valid():
                    cmr_ser.save()
                    return Response({"Message":"Update successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"Message":"Invalid Data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                cmr = CmrMashahir.objects.get(id=id)
                cmr.delete()
                return Response({"Message":"Delete successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class ContinerView(APIView):
    def get(self, request, id=None):
        
        if id:   
            try:
                paginator = PageNumberPagination()
                paginator.page_size = request.GET.get('page_size', 2)

                guide = Guide.objects.filter(company=id)
                guide_id = guide.values_list("id", flat=True)

                cont = Continer.objects.filter(guide__id__in=guide_id)\
                    .order_by('file_number')
                continer_id = cont.values_list("file_number", flat=True)
                print(continer_id)
                cont_page = paginator.paginate_queryset(cont, request) 
                current_page_continer_ids = [c.file_number for c in cont_page]
                cont_ser = ContinerSerializers(cont_page, many=True)
                guides = Guide.objects.filter(Q(continer__file_number__in=current_page_continer_ids), Q(is_valid=True))
                guide_ser = GuideSerializers(guides, many=True)

                booking = Booking.objects.filter(
                    guide__id__in=guide_id)
                booking_ser = BookingSerializers(booking, many=True)
                
                product = Product.objects.all()
                product_ser = Productserializers(product, many=True)
                
                document_type = GetDocumentType.objects.all()
                document_type_ser = GetDocumentTypeSerializers(document_type, many=True)
                
                carrier = Carrier.objects.all()
                carrier_ser = CarrierSerializers(carrier, many=True)
                
                route = Route.objects.all()
                route_ser = RouteSerializers(route, many=True)
                
                truck = Truck.objects.all()
                truck_ser = TruckSerializers(truck, many=True)
                
                driver = Driver.objects.all()
                driver_ser = DriverSerializers(driver, many=True)
                
                # bill = BillDescribtion.objects.filter(continer__file_number__in=continer_id)\
                #     .distinct('id')
                bill = BillDescribtion.objects.filter(
                    continer__file_number__in=current_page_continer_ids
                ).distinct()
                # bill_page = paginator.paginate_queryset(bill, request)
                bill_ser = BillDescribtionSeriallizer(bill, many=True)
                
                owner_of_good = OwnersOfGoods.objects.all()
                owner_of_good_ser = OwnersOfGoodsSeriallizer(owner_of_good, many=True)
                
                exit_continer = ExitContiner.objects.filter(
                    continer__file_number__in=current_page_continer_ids)
                exit_continer_ser = ExitContinerSerializers(exit_continer, many=True)
                
                ship = Sailing.objects.all()
                ship_ser = SailingSerializer(ship, many=True)
                
                return paginator.get_paginated_response({
                    "continer": cont_ser.data,
                    "product": product_ser.data, 
                    "document_type": document_type_ser.data,
                    "carrier": carrier_ser.data, 
                    "route": route_ser.data,
                    "truck": truck_ser.data, 
                    "driver": driver_ser.data,
                    "bill": bill_ser.data,
                    'booking': booking_ser.data,
                    'guide': guide_ser.data,
                    "owner_of_good": owner_of_good_ser.data,
                    "ship": ship_ser.data,
                    "exit_continer": exit_continer_ser.data
                })
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            try:
                paginator = PageNumberPagination()
                paginator.page_size = request.GET.get('page_size', 2)
                booking = Booking.objects.all()
                booking_ser = BookingSerializers(booking, many=True)

                guide = Guide.objects.all()
                guide_ser = GuideSerializers(guide, many=True)

                cont = Continer.objects.all()
                cont_page = paginator.paginate_queryset(cont, request)
                cont_ser = ContinerSerializers(cont_page, many=True)

                product  =Product.objects.all()
                product_ser = Productserializers(product, many=True)

                document_type = GetDocumentType.objects.all()
                document_type_ser = GetDocumentTypeSerializers(document_type, many=True)

                carrier = Carrier.objects.all()
                carrier_ser = CarrierSerializers(carrier, many=True)

                route = Route.objects.all()
                route_ser = RouteSerializers(route, many=True)

                truck = Truck.objects.all()
                truck_ser = TruckSerializers(truck, many=True)

                driver = Driver.objects.all()
                driver_ser = DriverSerializers(driver, many=True)

                bill = BillDescribtion.objects.all().order_by('-id')
                # bill_page = paginator.paginate_queryset(bill, request)
                bill_ser = BillDescribtionSeriallizer(bill, many=True)

                owner_of_good = OwnersOfGoods.objects.all()
                owner_of_good_ser = OwnersOfGoodsSeriallizer(owner_of_good, many=True)

                ship = Sailing.objects.all()
                ship_ser = SailingSerializer(ship, many=True)

                return paginator.get_paginated_response({"continer":cont_ser.data, "product":product_ser.data, 
                                "document_type":document_type_ser.data, "carrier":carrier_ser.data, 
                                "route":route_ser.data, "truck":truck_ser.data, 
                                "driver":driver_ser.data, "bill":bill_ser.data,
                                'booking':booking_ser.data,
                                'guide':guide_ser.data, "owner_of_good":owner_of_good_ser.data,
                                "ship":ship_ser.data},
                                )
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)

    def post(self, request):
        
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        bill = BillDescribtion.objects.get(id=request.data.get('bill'))
        if request.data.get('product') != None:
            product = Product.objects.get(id=request.data.get('product'))
        else:
            product=None
        
        if request.data.get('packet_type') != None:
            packet_type = PackagingType.objects.get(id=request.data.get('packet_type'))
        else:
            packet_type=None
        
        if request.data.get('carrier') != None:
            carrier = Carrier.objects.get(id=request.data.get('carrier'))
        else:
            carrier=None
        
        if request.data.get('route') != None:
            route = Route.objects.get(id=request.data.get('route'))
        else:
            route=None
        
        if request.data.get('truck') != None:
            truck = Truck.objects.get(id=request.data.get('truck'))
        else:
            truck=None

        if request.data.get('driver') != None:
            driver = Driver.objects.get(id=request.data.get('driver'))
        else:
            driver=None

        company = MainCompany.objects.get(id=request.data.get('company'))
        if permission.add_continer and permission.add_list_item:
            Date_receipt_documents_jalali_date = request.data.get('Date_receipt_documents')
            download_date_jalali_date = request.data.get('download_date')
            warehouse_receipt_date_jalali_date = request.data.get('warehouse_receipt_date')
            try:
                # year, month, day = map(int, Date_receipt_documents_jalali_date.split('-'))
                # Date_receipt_documents_gregorian_date = jdatetime.date(year, month, day).togregorian()
                # year, month, day = map(int, download_date_jalali_date.replace('/', '-').split('-'))
                # download_date_gregorian_date = jdatetime.date(year, month, day).togregorian()
                # print(download_date_gregorian_date)
                # year, month, day = map(int, warehouse_receipt_date_jalali_date.split('-'))
                # warehouse_receipt_date_gregorian_date = jdatetime.date(year, month, day).togregorian()
                cont = Continer.objects.create( 
                    number = request.data.get('number'),
                    invoice = request.data.get('invoice'),
                    size = request.data.get('size'),
                    Date_receipt_documents =  Date_receipt_documents_jalali_date,
                    bill = bill,
                    product = product,
                    product_count = request.data.get('product_count'),
                    packet_type = packet_type,
                    carrier = carrier,
                    national_id = request.data.get('national_id'),
                    download_date = download_date_jalali_date,
                    status = request.data.get('status'),
                    warehouse_receipt_date = warehouse_receipt_date_jalali_date ,
                    weight_documents =  request.data.get('weight_documents'),
                    weight_bascol = request.data.get('weight_bascol'),
                    difference_weight = request.data.get('difference_weight'),
                    excess_tonnage = request.data.get('excess_tonnage'),
                    serial_number = request.data.get('serial_number'),
                    route = route,
                    truck = truck,
                    commission = request.data.get('commission'),
                    fare_amount = request.data.get('fare_amount'),
                    advance_rent_amount = request.data.get('advance_rent_amount'),
                    driver = driver,
                    contract_number = request.data.get('contract_number'),
                    freeTime = request.data.get('freeTime'),
                    description = request.data.get('description'),
                    
                )
                cont.save()
                bill.containers_count = bill.containers_count + 1
                bill.save()
                sum_fee = SumFees.objects.create(
                    
                    sum_fees= 0,
                    continer = cont
                )
                sum_fee.save()
                guide = Guide.objects.create(
                continer=cont,
                guide_number=None,
                guide_id_barjame=None,
                company =company,
                is_active=None,
                sender=None,
                receiver=None,
                carrier=None,
                next_carrier=None,
                carrier_descriptions=None,
                place_date_delivery_goods=None,
                location_date_loading=None,
                attached_documents=None,
                sign_number=None,
                download_info=None,
                pure_agreements=None,
                sender_instructions=None,
                place_date_issue=None,
                date_place_receipt_goods=None,
                )
                guide.save()
                return Response({"Message":"Create successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)

    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_continer and permission.update_item:
            try:
                cont = Continer.objects.get(file_number=id)
                cont_ser = ContinerSerializers(cont, data=request.data,
                                    partial=True)
                if cont_ser.is_valid():
                    cont_ser.save()
                    return Response({"Message":"Update successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"Message":"Invalid Data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_continer and permission.delete_item:
            try:
                cont = Continer.objects.get(file_number=id)
                cont.delete()
                return Response({"Message":"Delete successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class CreateContinerView(APIView):
    def get(self, request, id):
        try:
            get_doct = GetDocumentType.objects.all()
            get_doct_ser = GetDocumentTypeSerializers(get_doct, many=True)

            product = Product.objects.all()
            product_ser = Productserializers(product, many=True)

            route = Route.objects.all()
            route_ser = RouteSerializers(route, many=True)

            truck = Truck.objects.all()
            truck_ser = TruckSerializers(truck, many=True)

            driver = Driver.objects.all()
            driver_ser = DriverSerializers(driver, many=True)

            ship = Sailing.objects.all()
            ship_ser = SailingSerializer(ship, many=True)

            good_owner = OwnersOfGoods.objects.all()
            good_owner_ser = OwnersOfGoodsSeriallizer(good_owner, many=True)

            demurrage_type = DemurrageType.objects.all()
            demurrage_type_ser = DemurrageTypeSerializers(demurrage_type, many=True)

            packet_type = PackagingType.objects.all()
            packet_type_ser = PackagingTypeSerializers(packet_type, many=True)

            carrier = Carrier.objects.all()
            carrier_ser = CarrierSerializers(carrier, many=True)

            return Response({"document_type":get_doct_ser.data,
                             "product":product_ser.data, 
                             "route":route_ser.data, 
                             "truck":truck_ser.data, 
                             "driver":driver_ser.data, 
                             "ship":ship_ser.data, 
                             "owner_of_good":good_owner_ser.data, 
                             "demurrage_type":demurrage_type_ser.data, 
                             "packet_type":packet_type_ser.data, 
                             "carrier":carrier_ser.data},
                             status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
class GuideView(APIView): # راهنامه
    def get(self, request, id):
        try:
            continer = Continer.objects.get(file_number=id)
            print(continer.bill.company.id)
            continer_ser = ContinerSerializers(continer, many=False)
            guide = Guide.objects.get(continer=continer.file_number)
            guide_ser = GuideSerializers(guide, many=False)
            truck = Truck.objects.get(id=continer.truck.id)
            truck_ser =TruckSerializers(truck, many=False)
            truck_list = Truck.objects.all()
            truck_list_ser = TruckSerializers(truck_list, many=True)
            carrier = Carrier.objects.all()
            carrier_ser = CarrierSerializers(carrier, many=True)
            route = Route.objects.all()
            route_ser = RouteSerializers(route, many=True)
            if continer.bill.company.id == 1:
                cmr = CmrBarjame.objects.all()
                cmr_ser = CmrBarjameSerializers(cmr, many=True)
            elif continer.bill.company.id == 2:
                cmr = CmrFerdows.objects.all()
                cmr_ser = CmrFerdowsserializers(cmr, many=True)
            elif continer.bill.company.id == 3:
                cmr = CmrSalamn.objects.all()
                cmr_ser = CmrSalamnSerializers(cmr, many=True)
            elif continer.bill.company.id == 4:
                cmr = CmrMashahir.objects.all()
                cmr_ser = CmrMashahirSerializers(cmr, many=True)
            
            volume = VolumeGuide.objects.filter(guide=guide)
            volume_ser = VolumeGuideSerializers(volume, many=True) 
            volum_sum = 0 
            for i in volume:
                volum_sum = volum_sum + i.volume 
            return Response({"guide":guide_ser.data, "continer":continer_ser.data,
                            "Truck":truck_ser.data, "cmr":cmr_ser.data,
                            "guide_volume":volume_ser.data, "truck_list":truck_list_ser.data,
                            "carrier":carrier_ser.data, "route":route_ser.data, "volum_sum":volum_sum}, 
                            status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_400_BAD_REQUEST)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                # guide_id = CmrFerdows.objects.get(id=request.data.get('guide_id'))
                continer = Continer.objects.get(file_number=request.data.get('continer'))
                company= MainCompany.objects.get(id=request.data.get('company'))
                if company.id == 1:
                    guide_id = CmrBarjame.objects.get(id=request.data.get('guide_id'))
                    guide = Guide.objects.create(
                    continer=continer,
                    guide_number=request.data.get("guide_number"),
                    guide_id_barjame=guide_id,
                    company =company,
                    is_active=request.data.get("is_active"),
                    sender=request.data.get("sender"),
                    receiver=request.data.get("receiver"),
                    carrier=request.data.get("carrier"),
                    next_carrier=request.data.get("next_carrier"),
                    carrier_descriptions=request.data.get("carrier_descriptions"),
                    place_date_delivery_goods=request.data.get("place_date_delivery_goods"),
                    location_date_loading=request.data.get("location_date_loading"),
                    attached_documents=request.data.get("attached_documents"),
                    sign_number=request.data.get("sign_number"),
                    download_info=request.data.get("download_info"),
                    pure_agreements=request.data.get("pure_agreements"),
                    sender_instructions=request.data.get("sender_instructions"),
                    place_date_issue=request.data.get("place_date_issue"),
                    date_place_receipt_goods=request.data.get("date_place_receipt_goods"),
                    )
                elif company.id == 2:
                    guide_id = CmrFerdows.objects.get(id=request.data.get('guide_id'))
                    guide = Guide.objects.create(
                    continer=continer,
                    guide_number=request.data.get("guide_number"),
                    guide_id_ferdows=guide_id,
                    company =company,
                    is_active=request.data.get("is_active"),
                    sender=request.data.get("sender"),
                    receiver=request.data.get("receiver"),
                    carrier=request.data.get("carrier"),
                    next_carrier=request.data.get("next_carrier"),
                    carrier_descriptions=request.data.get("carrier_descriptions"),
                    place_date_delivery_goods=request.data.get("place_date_delivery_goods"),
                    location_date_loading=request.data.get("location_date_loading"),
                    attached_documents=request.data.get("attached_documents"),
                    sign_number=request.data.get("sign_number"),
                    download_info=request.data.get("download_info"),
                    pure_agreements=request.data.get("pure_agreements"),
                    sender_instructions=request.data.get("sender_instructions"),
                    place_date_issue=request.data.get("place_date_issue"),
                    date_place_receipt_goods=request.data.get("date_place_receipt_goods"),
                    )
                elif company.id == 3:
                    guide_id = CmrSalamn.objects.get(id=request.data.get('guide_id'))
                    guide = Guide.objects.create(
                    continer=continer,
                    guide_number=request.data.get("guide_number"),
                    guide_id_salman=guide_id,
                    company =company,
                    is_active=request.data.get("is_active"),
                    sender=request.data.get("sender"),
                    receiver=request.data.get("receiver"),
                    carrier=request.data.get("carrier"),
                    next_carrier=request.data.get("next_carrier"),
                    carrier_descriptions=request.data.get("carrier_descriptions"),
                    place_date_delivery_goods=request.data.get("place_date_delivery_goods"),
                    location_date_loading=request.data.get("location_date_loading"),
                    attached_documents=request.data.get("attached_documents"),
                    sign_number=request.data.get("sign_number"),
                    download_info=request.data.get("download_info"),
                    pure_agreements=request.data.get("pure_agreements"),
                    sender_instructions=request.data.get("sender_instructions"),
                    place_date_issue=request.data.get("place_date_issue"),
                    date_place_receipt_goods=request.data.get("date_place_receipt_goods"),
                    )
                elif company.id == 4:
                   guide_id = CmrFerdows.objects.get(id=request.data.get('guide_id'))
                   guide = Guide.objects.create(
                    continer=continer,
                    guide_number=request.data.get("guide_number"),
                    guide_id_salman=guide_id,
                    company =company,
                    is_active=request.data.get("is_active"),
                    sender=request.data.get("sender"),
                    receiver=request.data.get("receiver"),
                    carrier=request.data.get("carrier"),
                    next_carrier=request.data.get("next_carrier"),
                    carrier_descriptions=request.data.get("carrier_descriptions"),
                    place_date_delivery_goods=request.data.get("place_date_delivery_goods"),
                    location_date_loading=request.data.get("location_date_loading"),
                    attached_documents=request.data.get("attached_documents"),
                    sign_number=request.data.get("sign_number"),
                    download_info=request.data.get("download_info"),
                    pure_agreements=request.data.get("pure_agreements"),
                    sender_instructions=request.data.get("sender_instructions"),
                    place_date_issue=request.data.get("place_date_issue"),
                    date_place_receipt_goods=request.data.get("date_place_receipt_goods"),
                    )
                else:
                    pass
                guide.save()
                return Response({"Message":"Create successfully"},
                                status=status.HTTP_201_CREATED)
            
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_list or permission.update_item:
            try:
                guide = Guide.objects.get(id=id)
                guide_ser = GuideSerializers(guide, data=request.data, partial=True)
                if guide_ser.is_valid():
                    guide_ser.save()
                    guide.is_valid = True
                    guide.save()
                    return Response({"Message":"Update successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    print(guide_ser.errors)
                    return Response({"Message":"Invalid Data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                guide = Guide.objects.get(id=id)
                guide.delete()
                return Response({"Message":"Delete successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_404_NOT_FOUND)

class GuidePluse(APIView): # get befor put
    def get(self, request, id):
        try:
            guide = Guide.objects.get(id=id)
            guide_ser = GuideSerializers(guide)
            return Response({"data":guide_ser.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
class VolumeGuideView(APIView):
    def get(self, request, id):
        try:
            guide =Guide.objects.get(id=id)
            volume = VolumeGuide.objects.filter(guide=guide)
            volume_ser = VolumeGuideSerializers(volume, many=True)
            return Response({"data":volume_ser.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message":"ERROR"}, status=status.HTTP_400_BAD_REQUEST)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)

        if permission.add_list_item:
            try:
                guide = Guide.objects.get(id=request.data.get('guide'))
                volume = VolumeGuide.objects.create(
                    guide = guide,
                    descriptions = request.data.get('descriptions'),
                    volume = request.data.get('volume')
                )
                volume.save()
                return Response({"Message":"Create successfuly"},
                                status=status.HTTP_201_CREATED)
            except Exception as e:
                print(e)
                return Response({"Message":"ERROR"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"Message":"You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_item or permission.delete_list:
            try:
                volume = VolumeGuide.objects.get(id=id)
                volume.delete()
                return Response({"Message":"Delete successfuly"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"ERROR"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"Message":"You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class DemurrageView(APIView):
    def get(self, request, id):
        try:
            demurrage = Demurrage.objects.filter(continer=id)
            demurrage_ser = DemurrageSerializers(demurrage, many=True)
            continer = Continer.objects.get(file_number=id)
            continer_ser = ContinerSerializers(continer, many=False)
            try: 
                    guide = Guide.objects.get(continer=continer)
                    if Booking.objects.get(guide=guide):
                        booking = Booking.objects.get(guide=guide)
                        booking_ser = BookingSerializers(booking, many=False)
                        return Response({"demurrage":demurrage_ser.data, 
                            "continer":continer_ser.data, "booking":booking_ser.data}, 
                            status=status.HTTP_200_OK)
                    else:
                        return Response({"demurrage":demurrage_ser.data, 
                            "continer":continer_ser.data}, 
                            status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"demurrage":demurrage_ser.data, 
                            "continer":continer_ser.data}, 
                            status=status.HTTP_200_OK)
        except Exception as e :
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_400_BAD_REQUEST)
        
    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                company = MainCompany.objects.get(id=request.data.get('company'))
                continer = Continer.objects.get(file_number=request.data.get('continer'))
                demurrage = Demurrage.objects.create(
                    demurrage_typpe = request.data.get('demurrage_typpe'),
                    type_info = request.data.get('type'),
                    date = request.data.get('date'),
                    price = request.data.get('price'),
                    description = request.data.get('description'),
                    file = request.data.get('file'),
                    company =company,
                    continer =continer
                )
                demurrage.save()
                return Response({"Message":"Create successfully"},
                                    status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_404_NOT_FOUND)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_list or permission.update_item:
            try:
                demurrage = Demurrage.objects.get(id=id)
                demurrage_ser = DemurrageSerializers(demurrage,
                                                data=request.data, partial=True)
                if demurrage_ser.is_valid():
                    demurrage_ser.save()
                    return Response({"Message":"Update successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    print(demurrage_ser.errors)
                    return Response({"Message":"Invalid data"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_404_NOT_FOUND)
    
    def delete(slef, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_list or permission.delete_item:
            try:
                demurrage = Demurrage.objects.get(id=id)
                demurrage.delete()
                return Response({"Message":"Delete successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        

class ExitContinerView(APIView):
    def get(self, request, id):
        try:
            # paginator = PageNumberPagination()
            # paginator.page_size = request.GET.get('page_size', 90)

            continer = Continer.objects.all()
            # continer_page = paginator.paginate_queryset(continer, request)
            continer_ser = continerPlusSerializers(continer, many=True)

            exit_containers = ExitContiner.objects.all()
            exit_continer_ser = ExitContinerSerializers(exit_containers, many=True)

            guide = Guide.objects.filter(company=id)
            guide_ser = GuideSerializers(guide, many=True)

            booking = Booking.objects.filter(company=id)
            booking_ser = BookingSerializers(booking, many=True)
            
            return Response({"continer":continer_ser.data,
                            "exit_continer":exit_continer_ser.data,
                            "guide":guide_ser.data,
                            "booking":booking_ser.data},
                            status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_400_BAD_REQUEST)

    def post(slef, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                cont = Continer.objects.get(file_number=request.data.get('id'))
                exit_cont = ExitContiner.objects.create(
                        continer = cont,
                        type = 'کانتینرهای خارج شده',
                        exit_date = request.data.get('exit_date'),  
                    )
                exit_cont.save()
                return Response({"Message":"Exit Continer successfully Created"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"you don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        try:
            exit_continer = ExitContiner.objects.get(id=id)
            exit_continer_ser = ExitContinerSerializers(exit_continer,
                                                data=request.data, partial=True)
            if exit_continer_ser.is_valid():
                exit_continer_ser.save()
                return Response({"Message":"Update successfully"},
                                status=status.HTTP_200_OK)
            else:
                return Response({"Message":"Invalid Data"},
                                status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_404_NOT_FOUND)
        
class MiscellaneousLoadView(APIView):
    def get(self, request, id):
        try:
            # paginator = PageNumberPagination()
            # paginator.page_size = request.GET.get('page_size', 90)
            miscellaneous_load = MiscellaneousLoad.objects.filter(main_company=id)
            # miscellaneous_load_page = paginator.paginate_queryset(miscellaneous_load, request)
            miscellaneous_load_ser = MiscellaneousLoadSerializers(miscellaneous_load, many=True)
            return Response({"data":miscellaneous_load_ser.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"your Target doesn't exist"},
                            status=status.HTTP_200_OK)

    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                continer = Continer.objects.get(id=request.data.get('continer'))
                company = MainCompany.objects.get(id=request.data.get('company'))
                des_company = MainCompany.objects.get(id=request.data.get('des_company'))
                miscellaneous_load = MiscellaneousLoad.objects.create(
                    main_company = company,
                    continer = continer,
                    des_company =des_company
                )
                miscellaneous_load.save()
                return Response({"Message":"Create successfully"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"your Target doesn't exist"},
                                status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Message":"You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class LoadingFeesView(APIView):
    def get(self, request, id):
        try:
            continer = Continer.objects.get(file_number=id)
            continer_ser = ContinerSerializers(continer, many=False)
            fees = LoadingFees.objects.filter(continer=continer)
            fees_ser = LoadingFeesSerializers(fees, many=True)
            sum_fees = 0
            for i in fees:
                sum_fees = sum_fees + i.fees
            guide = Guide.objects.get(continer=continer.file_number)
            booking = Booking.objects.get(guide=guide.id)
            booking_ser = BookingSerializers(booking, many=False)            
            cost = CostList.objects.all()
            cost_ser = CostListSerializers(cost, many=True)
            return Response({"sum_fees":sum_fees ,"fees":fees_ser.data ,"continer":continer_ser.data,
                            "booking" : booking_ser.data, "costList":cost_ser.data}, 
                            status=status.HTTP_200_OK)
            
        except Exception as e:
            print(e)
            return Response({"message":"ERROR"}, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.add_list_item:
            try:
                continer = Continer.objects.get(file_number=request.data.get('continer'))
                company = MainCompany.objects.get(id=request.data.get('company'))
                cost = CostList.objects.get(id=request.data.get('cost'))
                fees = LoadingFees.objects.create(
                    continer =continer,
                    company =company,
                    cost =cost,
                    date = request.data.get('date'),
                    fees = request.data.get('fees'),
                    description = request.data.get('description'),
                )
                fees.save()
                sum_fee = SumFees.objects.get(continer=fees.continer)
                sum_fee.sum_fees = sum_fee.sum_fees + fees.fees
                sum_fee.save()
                return Response({"Message":"Create successfuly"},
                                status=status.HTTP_201_CREATED)
            
            except Exception as e:
                print(e)
                return Response({"Message":"ERROR"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"Message":"You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def put(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.update_list or permission.update_item:
            try:
                fees = LoadingFees.objects.get(id=id)
                fees_ser = LoadingFeesSerializers(fees, data=request.data,
                                                partial=True)
                if fees_ser.is_valid():
                    sum_fee = SumFees.objects.get(continer=fees.continer)
                    sum_fee.sum_fees = sum_fee.sum_fees - fees.fees
                    sum_fee.save()
                    fees_ser.save()
                    sum_fee.sum_fees = sum_fee.sum_fees + fees.fees
                    sum_fee.save()
                    return Response({"Message":"Update successfuly"},
                                    status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"ERROR"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"Message":"You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
    def delete(self, request, id):
        token = request.data.get("token")
        user = get_user(token)
        permission = Permission.objects.get(user=user)
        if permission.delete_list or permission.delete_item:
            try:
                fees = LoadingFees.objects.get(id=id)
                sum_fee = SumFees.objects.get(continer=fees.continer)
                sum_fee.sum_fees = sum_fee.sum_fees - fees.fees
                sum_fee.save()
                fees.delete()
                return Response({"Message":"Delete successfuly"},
                                status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"Message":"ERROR"},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"Message":"You don't have permission"},
                            status=status.HTTP_403_FORBIDDEN)
        
class LogView(APIView):
    def get(self, request, id):
        try:
            logs = Log.objects.filter(Q(company=id) and Q(update_status=True))
            logs_ser = LogSerializers(logs, many=True)
            booking_id = logs.values_list("booking", flat=True).distinct()
            booking = Booking.objects.filter(Q(company=id) and Q(id__in=booking_id))
            
            booking_ser = BookingLogSerialzers(booking, many=True)
            return Response({"logs":logs_ser.data, "log_length":len(logs),
                            "Booking":booking_ser.data},
                            status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message":"ERROR"},
                            status=status.HTTP_400_BAD_REQUEST)
        
    def put(self, request):

        try:
            log_list = request.data.get('data')
            with transaction.atomic():
                print(f"log_list : {log_list}")
                for i in log_list:
                    print(i)
                    log = Log.objects.get(id=i)
                    log.update_status = False
                    log.status = 'دیده شده'
                    log.save()
                    # log_ser = LogSerializers(log, data=i, partial=True,
                    #             )
                    # print(log_ser)
                    # if log_ser.is_valid():
                    #     log_ser.save()          
                    # else:
                    #     print(log_ser.errors)
                    #     return Response({"Message":"ERROR"},
                    #                     status=status.HTTP_400_BAD_REQUEST)
                    
                return Response({"Message":"Update successfuly"},
                                    status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"Message":"ERROR"},
                            status=status.HTTP_400_BAD_REQUEST)
        
        
class GuideReportView(APIView):
    def get(self, request):
        try:
            guide = Guide.objects.all()
            guides_salman = []
            guides_barjaame = []
            guides_ferdows = []
            for i in guide:
                if i.company == 1:
                    guides_barjaame.append(i)
                    guide_ser_barjame = GuideReportSerialzers(guide, many=True)
                elif i.company == 2:
                    guides_ferdows.append(i)
                    guide_ser_ferdow = GuideReportSerialzers(guide, many=True)
                elif i.company == 3:
                    guides_salman.append(i)
                    guide_ser_salman = GuideReportSerialzers(guide, many=True)
                elif i.company == 4:
                    pass
                else:
                    pass
            
            return Response({"guides_salman":guide_ser_salman.data,
                            "guides_ferdows":guide_ser_ferdow.data, 
                            "guides_barjame":guide_ser_barjame.data}, 
                            status=status.HTTP_200_OK)
                
        except Exception as e:
            print(e)
            return Response({"message":"ERROR"},
                            status=status.HTTP_400_BAD_REQUEST)
        

class SumFeesView(APIView):
    def get(self, request, id):
        try:
            sum_fee = SumFees.objects.filter(company=id)
            sum_fee_ser = SumFeesSerialzers(sum_fee, many=True)
            return Response({"data":sum_fee_ser.data},status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message":"ERROR"},
                            status=status.HTTP_400_BAD_REQUEST)
        
class ContinerBookingView(APIView): # get continer continer-booking page and continer-guide page
    def get(self, request,id):
        try:
            if id:
                paginator = PageNumberPagination()
                paginator.page_size = request.GET.get('page_size', 15)

                continer_booking = Guide.objects.filter(company=id)
                continer_booking_page = paginator.paginate_queryset(continer_booking, request) 
                serializer = ContinerBookingSerializers(continer_booking_page, many=True)

                route = Route.objects.all()
                route_ser = RouteSerializers(route, many=True)

                ship = Sailing.objects.all()
                ship_ser = SailingSerializer(ship, many=True)

                owner_goods = OwnersOfGoods.objects.all()
                owner_goods_ser = OwnersOfGoodsSeriallizer(owner_goods, many=True)

                carrier = Carrier.objects.all()
                carrier_ser = CarrierSerializers(carrier, many=True)

                guide = Guide.objects.filter(Q(is_valid=True) & Q(company=id))
                guide2 = Guide.objects.filter(Q(is_valid=False) & Q(company=id))
                guide3 = Guide.objects.filter(Q(is_active=False) & Q(company=id))
                guide_id = guide.values_list("id", flat=True)

                booking = Booking.objects.filter(guide__id__in=guide_id)

                sum_fee = SumFees.objects.filter(Q(sum_fees__gt=0) & Q(continer__bill__company=id))

                demurrage = Demurrage.objects.filter(Q(continer__bill__company=id))
                demurrage_id = demurrage.values_list("continer__file_number", flat=True)

                exit_cont = len(ExitContiner.objects.filter(Q(types='کانتینرهای خارج شده') 
                                                            & Q(continer__bill__company=id)))
                
                exit_cont2 = len(ExitContiner.objects.filter(Q(types='کانتینرهای مهلت خروج گذشته') 
                                                            & Q(continer__bill__company=id)))
                
                exit_cont3 = len(ExitContiner.objects.filter(Q(types='کانتینرهای خارج نشده')
                                                            & Q(continer__bill__company=id)))
                
                continer = len(Continer.objects.filter(Q(bill__company=id) & Q(download_date=None)))
                
                # guide_len = کل راهنامه ها و کانتینر دارای راهنامه
                # booking_len =  کل راهنامه ها ی فاقد بوکینگ
                # booking =  کل کانتینر های دارای بوکینگ
                # cont_without_booking = کانتینر های فاقد بوکینگ
                # guide2 = کانتینر فاقد راهنامه
                # guide3 = تعداد راهنامه باطل شده
                #sum_fee = کانتینر دارای هزینه
                # not_fee = کانتینر فاقد هزینه
                # demurrage = کانتینر دارای هزینه دمراژ
                # not_demurrage = کانتینر فاقد هزینه دمراژ
                # exit_cont = تعداد کانتینر خارج شده
                # exit_cont3 = تعداد کانتینر خارج نشده
                # exit_cont2 = تعداد کانتینر مهلت خروج گذشته
                # not_download = فاقد بارگیری
                return paginator.get_paginated_response({"data":serializer.data, "route":route_ser.data,
                                            "ship":ship_ser.data, "owner_goods":owner_goods_ser.data,
                                            "carrier":carrier_ser.data, "guide_len":len(guide), 
                                            "booking_len":len(guide)- len(booking),
                                            "booking":len(booking),
                                            "cont_without_booking":len(continer_booking) -len(booking),
                                            "guide2":len(guide2), "guide3":len(guide3),  
                                            "sum_fee":len(sum_fee), 
                                            "not_fee":len(continer_booking) - len(sum_fee),
                                            "demurrage":len(set(demurrage_id)),
                                            "not_demurrage":len(continer_booking) - len(set(demurrage_id)),
                                            "exit_cont":exit_cont, "exit_cont2":exit_cont2,
                                            "exit_cont3":exit_cont3, 
                                            "not_download":continer
                                           })
        
        
        except Exception as e:
            print(e)
            return Response({"message":"ERROR"},
                            status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        try:
            paginator = PageNumberPagination()
            paginator.page_size = request.GET.get('page_size', 20)
            
            serial_number = request.data.get("serial_number")
            number = request.data.get("number")
            contract_number = str(request.data.get("contract_number")) \
                if request.data.get('contract_number') else None
            truck = request.data.get("truck")
            route = request.data.get("route")
            carrier = request.data.get("carrier")
            bill = request.data.get("bill")
            file_number = request.data.get("file_number")
            owner_goods = request.data.get("owner_goods")
            ship = request.data.get("ship")
            booking_num_start = request.data.get("booking_num_start")
            booking_num_end = request.data.get("booking_num_end")
            start_date_str = request.data.get("start_date")
            end_date_str = request.data.get("end_date")
            start_guide_date_str = request.data.get("start_guide_date")
            end_guide_date_str = request.data.get("end_guide_date")
            start_guide_num = request.data.get("start_guide_num")
            end_guide_num = request.data.get("end_guide_num")
            cont_status = request.data.get("status")
            company = request.data.get("company")

            query = Q()
            bill_query = Q()
            guide_query = Q()
            if serial_number:
                query &= Q(serial_number__exact=serial_number)

            if number:
                query &= Q(number__exact=number)
            
            if contract_number:
                query &= Q(contract_number__exact=contract_number)
            
            if truck:
                query &= Q(truck__License_plate_number__exact=truck)
            
            if route:
                query &= Q(route__title__exact=route)
            
            if carrier:
                query &= Q(carrier__name__exact=carrier)

            if bill:
                query &= Q(bill__bill_lading_number__exact=bill)

            if cont_status:
                query &= Q(status__exact=cont_status)

            if file_number:
                query &= Q(file_number__exact=file_number)

            if start_date_str and end_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                    query &= Q(download_date__range=(start_date, end_date))
                except ValueError as e:
                    return Response(
                        {"message": "Invalid date format. Use YYYY-MM-DD"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                    
            if owner_goods:
                query &= Q(bill__Owners_of_goods__name__exact=owner_goods)
            
            if ship:
                query &= Q(bill__ship__name__exact=ship)

            if booking_num_end and booking_num_start :
                bill_query &= Q(booking_number__range=(booking_num_start, booking_num_end))

            if start_guide_num and end_guide_num :
                guide_query &=  Q(guide_number__range=(start_guide_num, end_guide_num)) & Q(is_valid=True)
               
            if start_guide_date_str and end_guide_date_str:
                try:
                    start_date = datetime.strptime(start_guide_date_str, "%Y-%m-%d").date()
                    end_date = datetime.strptime(end_guide_date_str, "%Y-%m-%d").date()
                    guide_query &= Q(confirm_date__range=(start_date, end_date)) & Q(is_valid=True)
                except ValueError as e:
                    return Response(
                        {"message": "Invalid date format. Use YYYY-MM-DD"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
            print(request.data)
            continer = Continer.objects.filter(query)
            continer_id = continer.values_list("file_number", flat=True)
            booking = Booking.objects.filter(bill_query
                )
            booking_id = booking.values_list("guide_id", flat=True)
            print(continer_id)
            
            if serial_number or number or contract_number or truck or route or bill \
                or file_number or start_date_str or end_date_str or owner_goods or \
                ship or cont_status or carrier:
                guide_query &= Q(continer__file_number__in=continer_id) 
                print(f'guide_queryccccccccccccccccccccccccc {guide_query}')

            if start_guide_num or end_guide_num  or start_guide_date_str or end_guide_date_str:
                print('asdf')
                guide_query &= Q(continer__file_number__in=continer_id) & Q(is_valid=True)

            if  booking_num_start and booking_num_end :
                print('sfadsfds')
                guide_query &= Q(id__in=booking_id)
                
            print(f'guide_query {guide_query}')    
            guides = Guide.objects.filter(guide_query)
            print(f'guides {guides}')
            print(file_number, '======', query)
            continer_booking_page = paginator.paginate_queryset(guides, request) 
            ser = ContinerBookingSerializers(continer_booking_page, many=True)

            route = Route.objects.all()
            route_ser = RouteSerializers(route, many=True)

            ship = Sailing.objects.all()
            ship_ser = SailingSerializer(ship, many=True)

            carrier = Carrier.objects.all()
            carrier_ser = CarrierSerializers(carrier, many=True)

            owner_goods = OwnersOfGoods.objects.all()
            owner_goods_ser = OwnersOfGoodsSeriallizer(owner_goods, many=True)
            # ser = BookingSerializers(booking, many=True)
            return paginator.get_paginated_response({"data":ser.data,"route":route_ser.data,
                                            "ship":ship_ser.data, "owner_goods":owner_goods_ser.data, 
                                            "carrier": carrier_ser.data
                                            },
                        )
        
        except Exception as e:
            print(e)
            return Response({"message":"ERROR"},
                            status=status.HTTP_400_BAD_REQUEST)
        
class GenerateReportPDF(APIView):
    """
    A class-based view using DRF's APIView to handle PDF generation.
    Receives POST data, populates an HTML template, and returns a 
    downloadable PDF file.
    """
    def post(self, request):
        # 1. Get data from the request. DRF's `request.data` handles JSON, form data, etc.
        data = request.data
        context = {
            'report_date': data.get('report_date', ''),
            'booking_number': data.get('booking_number', ''),
            'issue_date': data.get('issue_date', ''),
            'waybill_number': data.get('waybill_number', ''),
            'carrier_company': data.get('carrier_company', ''),
            'loading_date': data.get('loading_date', ''),
            'route': data.get('route', ''),
            'truck_number': data.get('truck_number', ''),
            'size': data.get('size', ''),
            'cargo_weight': data.get('cargo_weight', ''),
            'driver_name': data.get('driver_name', ''),
            'contract_number': data.get('contract_number', ''),
            'shipper_name': data.get('shipper_name', ''),
            'shipper_commission_percentage': data.get('shipper_commission_percentage', ''),
            'freight_cost': data.get('freight_cost', ''),
            'shipper_commission_amount': data.get('shipper_commission_amount', ''),
            'excess_tonnage_quantity_shipper': data.get('excess_tonnage_quantity_shipper', '0'),
            'excess_tonnage_amount_shipper': data.get('excess_tonnage_amount_shipper', '0'),
            'excess_tonnage_weight_shipper': data.get('excess_tonnage_weight_shipper', '0'),
            'shipper_total': data.get('shipper_total', ''),
            'company_name': data.get('company_name', ''),
            'company_commission': data.get('company_commission', ''),
            'waybill_insurance': data.get('waybill_insurance', ''),
            'intermediary_commission': data.get('intermediary_commission', '0'),
            'company_total': data.get('company_total', ''),
            'driver_balance': data.get('driver_balance', ''),
            'difference': data.get('difference', '0'),
            'notes': data.get('notes', ''),
        }

        # 2. Load and render the HTML template
        template = get_template('form.html')
        html_string = template.render(context)

        # 3. Generate the PDF using WeasyPrint
        pdf_file = weasyprint.HTML(string=html_string).write_pdf()

        # 4. Create a standard Django HttpResponse to serve the file.
        #    This is the correct approach for file downloads even within an APIView.
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report.pdf"'
        
        return response